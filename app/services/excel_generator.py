import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage

from app.core.config import Settings
from app.core.errors import AppError
from app.domain.report_models import DailyReport, WorkItem
from app.domain.template_models import TemplateProfile


class ExcelGenerator:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.settings.ensure_directories()

    def generate(
        self,
        template_path: Path,
        profile: TemplateProfile,
        report: DailyReport,
        photos_by_id: dict[str, Path],
    ) -> Path:
        output_dir = self.settings.temporary_outputs_dir / report.report_id
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"Daily_Report_{report.report_date.replace('-', '')}.xlsx"
        shutil.copy2(template_path, output_path)

        workbook = load_workbook(output_path)
        for field_name, mapping in profile.fields.items():
            if mapping.sheet_name not in workbook.sheetnames:
                raise AppError(f"模板字段 {field_name} 引用了不存在的工作表：{mapping.sheet_name}")
            workbook[mapping.sheet_name][mapping.cell] = self._field_value(field_name, report)

        for slot in profile.photo_slots:
            sheet = workbook[slot.sheet_name]
            for analysis in report.photo_analyses:
                if slot.accepted_photo_types and analysis.photo_type not in slot.accepted_photo_types:
                    continue
                photo_path = photos_by_id.get(analysis.photo_id)
                if not photo_path:
                    continue
                image = ExcelImage(str(photo_path))
                image.width = slot.width_px
                image.height = slot.height_px
                sheet.add_image(image, slot.anchor_cell)
                break

        workbook.save(output_path)
        return output_path

    def _field_value(self, field_name: str, report: DailyReport) -> str:
        values = {
            "report_date": report.report_date,
            "project_name": report.project_name,
            "departure_time": report.timeline.departure_time or "",
            "arrival_site_time": report.timeline.arrival_site_time or "",
            "leave_site_time": report.timeline.leave_site_time or "",
            "arrival_hotel_time": report.timeline.arrival_hotel_time or "",
            "work_content_chinese": self._numbered_work_items(report.work_items, "chinese"),
            "work_content_english": self._numbered_work_items(report.work_items, "english"),
            "next_day_plan_chinese": report.next_day_plan.polished_chinese,
            "next_day_plan_english": report.next_day_plan.english,
            "remarks": report.remarks,
        }
        return values.get(field_name, "")

    def _numbered_work_items(self, work_items: list[WorkItem], language: str) -> str:
        lines = []
        for index, item in enumerate(work_items, start=1):
            text = item.content.english if language == "english" else item.content.polished_chinese
            lines.append(f"{index}. {text}")
        return "\n".join(lines)
