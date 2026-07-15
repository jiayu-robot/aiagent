from pathlib import Path

from openpyxl import load_workbook

from app.domain.template_models import NonEmptyCell, TemplateInspection


def inspect_template(path: Path) -> TemplateInspection:
    workbook = load_workbook(path)
    non_empty_cells: list[NonEmptyCell] = []
    merged_ranges: dict[str, list[str]] = {}
    image_counts: dict[str, int] = {}

    for sheet in workbook.worksheets:
        merged_ranges[sheet.title] = [str(cell_range) for cell_range in sheet.merged_cells.ranges]
        image_counts[sheet.title] = len(getattr(sheet, "_images", []))
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    non_empty_cells.append(
                        NonEmptyCell(sheet_name=sheet.title, cell=cell.coordinate, value=str(cell.value))
                    )

    return TemplateInspection(
        sheets=workbook.sheetnames,
        non_empty_cells=non_empty_cells,
        merged_ranges=merged_ranges,
        image_counts=image_counts,
    )
