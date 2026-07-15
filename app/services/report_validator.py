from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.core.errors import AppError
from app.domain.report_models import DailyReport
from app.domain.template_models import TemplateProfile


REQUIRED_PROFILE_FIELDS = {
    "report_date",
    "project_name",
    "work_content_chinese",
    "work_content_english",
}


class ReportValidator:
    def validate(self, report: DailyReport, profile: TemplateProfile, template_path: Path) -> list[str]:
        self._validate_report_date(report.report_date)
        self._validate_work_items(report)
        self._validate_time_order(report)
        self._validate_related_photo_ids(report)
        self._validate_profile(profile, template_path)
        warnings = self._collect_photo_warnings(report)
        report.warnings = list(dict.fromkeys([*report.warnings, *warnings]))
        return report.warnings

    def _validate_report_date(self, value: str) -> None:
        try:
            datetime.strptime(value, "%Y-%m-%d")
        except ValueError as exc:
            raise AppError("日报日期格式必须是 YYYY-MM-DD") from exc

    def _validate_work_items(self, report: DailyReport) -> None:
        if not report.work_items:
            raise AppError("工作内容不能为空")

    def _validate_time_order(self, report: DailyReport) -> None:
        ordered_values = [
            report.timeline.departure_time,
            report.timeline.arrival_site_time,
            report.timeline.leave_site_time,
            report.timeline.arrival_hotel_time,
        ]
        parsed: list[tuple[int, datetime]] = []
        for index, value in enumerate(ordered_values):
            if value:
                try:
                    parsed.append((index, self._parse_time(value)))
                except ValueError as exc:
                    raise AppError("时间格式必须是 HH:MM 或 HH:MM:SS") from exc
        for (_, previous), (_, current) in zip(parsed, parsed[1:]):
            if previous > current:
                raise AppError("时间顺序不正确：出发、到达现场、离开现场、到达酒店必须依次递增")

    def _parse_time(self, value: str) -> datetime:
        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
        raise ValueError(value)

    def _validate_related_photo_ids(self, report: DailyReport) -> None:
        existing = {analysis.photo_id for analysis in report.photo_analyses}
        for item in report.work_items:
            for photo_id in item.related_photo_ids:
                if photo_id not in existing:
                    raise AppError(f"工作事项关联的照片不存在：{photo_id}")

    def _validate_profile(self, profile: TemplateProfile, template_path: Path) -> None:
        missing = sorted(REQUIRED_PROFILE_FIELDS - set(profile.fields))
        if missing:
            raise AppError(f"缺少必需模板字段：{', '.join(missing)}")
        workbook = load_workbook(template_path)
        for field_name, mapping in profile.fields.items():
            if mapping.sheet_name not in workbook.sheetnames:
                raise AppError(f"模板字段 {field_name} 引用了不存在的工作表：{mapping.sheet_name}")
        for slot in profile.photo_slots:
            if slot.sheet_name not in workbook.sheetnames:
                raise AppError(f"照片插槽引用了不存在的工作表：{slot.sheet_name}")

    def _collect_photo_warnings(self, report: DailyReport) -> list[str]:
        warnings: list[str] = []
        for analysis in report.photo_analyses:
            warnings.extend(analysis.warnings)
            if analysis.confidence < 0.6:
                warnings.append(f"照片 {analysis.original_filename} 分析置信度较低，需要人工确认。")
            if analysis.exif_datetime and analysis.visible_time:
                exif_time = analysis.exif_datetime.split(" ", 1)[1][:5]
                if exif_time != analysis.visible_time:
                    warnings.append(f"照片 {analysis.original_filename} 的 EXIF 时间与画面时间不一致，需要人工确认。")
        return warnings
