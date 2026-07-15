from pathlib import Path

from openpyxl import Workbook

from app.services.template_inspector import inspect_template


def create_inspection_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily"
    sheet["A1"] = "项目"
    sheet["B2"] = "日期"
    sheet.merge_cells("C3:D3")
    workbook.create_sheet("Photos")
    workbook.save(path)


def test_inspect_template_lists_sheets_cells_and_merged_ranges(tmp_path):
    template_path = tmp_path / "template.xlsx"
    create_inspection_workbook(template_path)

    inspection = inspect_template(template_path)

    assert inspection.sheets == ["Daily", "Photos"]
    assert {cell.cell for cell in inspection.non_empty_cells} == {"A1", "B2"}
    assert inspection.merged_ranges["Daily"] == ["C3:D3"]
    assert inspection.image_counts == {"Daily": 0, "Photos": 0}
