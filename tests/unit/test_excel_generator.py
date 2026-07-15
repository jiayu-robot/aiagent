from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image

from app.core.config import Settings
from app.domain.photo_models import PhotoAnalysis
from app.domain.report_models import DailyReport, MultilingualText, Timeline, WorkItem
from app.domain.template_models import PhotoSlot, TemplateFieldMapping, TemplateProfile
from app.services.excel_generator import ExcelGenerator


def create_template(path: Path) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily"
    sheet["A1"] = "DO NOT CHANGE"
    workbook.save(path)
    return path.read_bytes()


def create_photo(path: Path) -> None:
    Image.new("RGB", (80, 60), color="green").save(path, format="JPEG")


def sample_report() -> DailyReport:
    return DailyReport(
        report_id="report-1",
        report_date="2026-07-15",
        project_name="现场工程项目",
        timeline=Timeline(departure_time="08:00", arrival_site_time="09:00"),
        work_items=[
            WorkItem(
                title_chinese="主控检查",
                title_english="Main Controller Inspection",
                content=MultilingualText(
                    raw_chinese="检查主控",
                    polished_chinese="检查主控系统的软件版本。",
                    english="Checked the software version of the main controller system.",
                ),
                related_photo_ids=["photo-1"],
            )
        ],
        next_day_plan=MultilingualText(
            raw_chinese="继续测试",
            polished_chinese="继续开展系统测试。",
            english="Continue system testing.",
        ),
        photo_analyses=[
            PhotoAnalysis(
                photo_id="photo-1",
                original_filename="site.jpg",
                photo_type="工作现场",
                description="site",
                confidence=0.9,
            )
        ],
    )


def sample_profile() -> TemplateProfile:
    return TemplateProfile(
        template_filename="template.xlsx",
        fields={
            "report_date": TemplateFieldMapping(sheet_name="Daily", cell="B2"),
            "project_name": TemplateFieldMapping(sheet_name="Daily", cell="B3"),
            "departure_time": TemplateFieldMapping(sheet_name="Daily", cell="B4"),
            "arrival_site_time": TemplateFieldMapping(sheet_name="Daily", cell="B5"),
            "work_content_chinese": TemplateFieldMapping(sheet_name="Daily", cell="B8", language="chinese"),
            "work_content_english": TemplateFieldMapping(sheet_name="Daily", cell="B9", language="english"),
            "next_day_plan_chinese": TemplateFieldMapping(sheet_name="Daily", cell="B10", language="chinese"),
            "next_day_plan_english": TemplateFieldMapping(sheet_name="Daily", cell="B11", language="english"),
        },
        photo_slots=[
            PhotoSlot(
                sheet_name="Daily",
                anchor_cell="A13",
                width_px=160,
                height_px=120,
                accepted_photo_types=["工作现场"],
            )
        ],
    )


def test_excel_generator_writes_cells_in_copy_and_preserves_template(tmp_path):
    settings = Settings(app_data_root=tmp_path)
    template = tmp_path / "template.xlsx"
    original_bytes = create_template(template)
    photo_path = tmp_path / "photo.jpg"
    create_photo(photo_path)

    output = ExcelGenerator(settings).generate(
        template_path=template,
        profile=sample_profile(),
        report=sample_report(),
        photos_by_id={"photo-1": photo_path},
    )

    workbook = load_workbook(output)
    sheet = workbook["Daily"]
    assert sheet["A1"].value == "DO NOT CHANGE"
    assert sheet["B2"].value == "2026-07-15"
    assert sheet["B8"].value.startswith("1. 检查主控系统")
    assert sheet["B9"].value.startswith("1. Checked the software")
    assert len(sheet._images) == 1
    assert template.read_bytes() == original_bytes


def test_excel_generator_supports_template_specific_values_and_time_cells(tmp_path):
    settings = Settings(app_data_root=tmp_path)
    template = tmp_path / "template.xlsx"
    create_template(template)
    report = sample_report()
    report.template_values = {
        "engineer_1_name": "Mengyang Li",
        "engineer_2_name": "Jiayu Yang",
        "engineer_1_check_in": "08:44:00",
        "engineer_1_check_out": "15:39:00",
    }
    profile = TemplateProfile(
        template_filename="template.xlsx",
        fields={
            "report_date": TemplateFieldMapping(sheet_name="Daily", cell="B2"),
            "project_name": TemplateFieldMapping(sheet_name="Daily", cell="B3"),
            "work_content_chinese": TemplateFieldMapping(sheet_name="Daily", cell="B8", language="chinese"),
            "work_content_english": TemplateFieldMapping(sheet_name="Daily", cell="B9", language="english"),
            "engineer_1_name": TemplateFieldMapping(sheet_name="Daily", cell="A9"),
            "engineer_2_name": TemplateFieldMapping(sheet_name="Daily", cell="A10"),
            "engineer_1_check_in": TemplateFieldMapping(sheet_name="Daily", cell="C9"),
            "engineer_1_check_out": TemplateFieldMapping(sheet_name="Daily", cell="D9"),
        },
        photo_slots=[],
    )

    output = ExcelGenerator(settings).generate(template, profile, report, {})

    sheet = load_workbook(output)["Daily"]
    assert sheet["A10"].value == "Jiayu Yang"
    assert sheet["C9"].value.hour == 8
    assert sheet["C9"].value.minute == 44
    assert sheet["D9"].value.hour == 15
    assert sheet["D9"].value.minute == 39


def test_excel_generator_allows_template_value_override_for_standard_field(tmp_path):
    settings = Settings(app_data_root=tmp_path)
    template = tmp_path / "template.xlsx"
    create_template(template)
    report = sample_report()
    report.template_values = {"report_date": "15/07/2026 Wednesday"}
    profile = TemplateProfile(
        template_filename="template.xlsx",
        fields={
            "report_date": TemplateFieldMapping(sheet_name="Daily", cell="B2"),
            "project_name": TemplateFieldMapping(sheet_name="Daily", cell="B3"),
            "work_content_chinese": TemplateFieldMapping(sheet_name="Daily", cell="B8", language="chinese"),
            "work_content_english": TemplateFieldMapping(sheet_name="Daily", cell="B9", language="english"),
        },
        photo_slots=[],
    )

    output = ExcelGenerator(settings).generate(template, profile, report, {})

    assert load_workbook(output)["Daily"]["B2"].value == "15/07/2026 Wednesday"


def test_excel_generator_writes_to_top_left_when_mapping_points_inside_merged_range(tmp_path):
    settings = Settings(app_data_root=tmp_path)
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily"
    sheet.merge_cells("A42:H43")
    sheet["A42"] = "Remarks:"
    workbook.save(template)
    report = sample_report()
    report.template_values = {"remarks_detail": "Remarks:\nDaily updated report."}
    profile = TemplateProfile(
        template_filename="template.xlsx",
        fields={
            "report_date": TemplateFieldMapping(sheet_name="Daily", cell="B2"),
            "project_name": TemplateFieldMapping(sheet_name="Daily", cell="B3"),
            "work_content_chinese": TemplateFieldMapping(sheet_name="Daily", cell="B8", language="chinese"),
            "work_content_english": TemplateFieldMapping(sheet_name="Daily", cell="B9", language="english"),
            "remarks_detail": TemplateFieldMapping(sheet_name="Daily", cell="B42"),
        },
        photo_slots=[],
    )

    output = ExcelGenerator(settings).generate(template, profile, report, {})

    sheet = load_workbook(output)["Daily"]
    assert sheet["A42"].value == "Remarks:\nDaily updated report."
