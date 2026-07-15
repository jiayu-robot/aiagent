from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image

from app.core.config import Settings
from app.main import create_app


def workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily"
    sheet["A1"] = "AI Daily Report"
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def image_bytes(color: str) -> bytes:
    image = Image.new("RGB", (60, 40), color=color)
    stream = BytesIO()
    image.save(stream, format="JPEG")
    return stream.getvalue()


def profile_payload() -> dict:
    return {
        "template_filename": "template.xlsx",
        "fields": {
            "report_date": {"sheet_name": "Daily", "cell": "B2", "language": "neutral", "optional": False},
            "project_name": {"sheet_name": "Daily", "cell": "B3", "language": "neutral", "optional": False},
            "departure_time": {"sheet_name": "Daily", "cell": "B4", "language": "neutral", "optional": True},
            "arrival_site_time": {"sheet_name": "Daily", "cell": "B5", "language": "neutral", "optional": True},
            "work_content_chinese": {"sheet_name": "Daily", "cell": "B8", "language": "chinese", "optional": False},
            "work_content_english": {"sheet_name": "Daily", "cell": "B9", "language": "english", "optional": False},
            "next_day_plan_chinese": {"sheet_name": "Daily", "cell": "B10", "language": "chinese", "optional": True},
            "next_day_plan_english": {"sheet_name": "Daily", "cell": "B11", "language": "english", "optional": True},
            "remarks": {"sheet_name": "Daily", "cell": "B12", "language": "neutral", "optional": True},
        },
        "photo_slots": [
            {
                "sheet_name": "Daily",
                "anchor_cell": "A14",
                "width_px": 160,
                "height_px": 120,
                "accepted_photo_types": ["工作现场", "出发打卡"],
            }
        ],
    }


def test_full_api_report_generation_flow(tmp_path):
    app = create_app(Settings(app_data_root=tmp_path))
    client = TestClient(app)

    upload_response = client.post(
        "/api/templates/upload",
        files={"file": ("template.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert upload_response.status_code == 200
    assert upload_response.json()["inspection"]["sheets"] == ["Daily"]

    profile_response = client.post("/api/templates/profile", json=profile_payload())
    assert profile_response.status_code == 200
    assert profile_response.json()["status"] == "saved"

    analyze_response = client.post(
        "/api/reports/analyze",
        data={"raw_chinese": "检查了一下主控，发现版本不一样，重新升级后正常了。"},
        files=[
            ("photos", ("departure.jpg", image_bytes("blue"), "image/jpeg")),
            ("photos", ("site.jpg", image_bytes("green"), "image/jpeg")),
        ],
    )
    assert analyze_response.status_code == 200
    report_id = analyze_response.json()["report_id"]
    assert analyze_response.json()["report"]["work_items"][0]["content"]["english"]
    assert len(analyze_response.json()["report"]["photo_analyses"]) == 2

    preview_response = client.get(f"/api/reports/{report_id}/preview")
    assert preview_response.status_code == 200
    preview = preview_response.json()["report"]
    preview["remarks"] = "现场已确认。"

    generate_response = client.post(f"/api/reports/{report_id}/generate", json={"report": preview})
    assert generate_response.status_code == 200
    download_url = generate_response.json()["download_url"]

    download_response = client.get(download_url)
    assert download_response.status_code == 200
    output_path = tmp_path / "downloaded.xlsx"
    output_path.write_bytes(download_response.content)
    workbook = load_workbook(output_path)
    sheet = workbook["Daily"]
    assert sheet["B3"].value == "现场工程项目"
    assert "Main Controller" in sheet["B9"].value
    assert sheet["B12"].value == "现场已确认。"
    assert len(sheet._images) == 1
